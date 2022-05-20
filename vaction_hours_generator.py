#!/usr/bin/env python3

'''
vacation_hours_generator.py

Takes a CSV from lightspeed retail POS, pulls out employee names and total hours worked to calculate how much
vacation time each employee has available.

-Written by Patrick Coyle

USAGE:

- On first run of this script, two directories will be created in the same directory as this
  file: to_parse and parsed_data

- REQUIRED DATA: Sales by line exported to .csv files in Lightspeed Retail for the month
  you wish to parse.

  These required files can be found the following way
  1. Go to Reports
  2. Go to Total Hours
  3. Change the date range to the to all those you wish to scan
  4. Pick "All Employees" from the drop down menu to the right of the dates
  5. Choose "All locations" from the drop down menu to the right of the employee selection
  6. Select "export" and save that csv to be put into the "to_parse" directory
  
'''

#TO DO

import csv, os, os.path, shutil

from datetime import datetime
from openpyxl import load_workbook

# csv - reads the data from Lightspeed Retail POS
# os - used for creating folders and checking folder contents
# os.path - Used for checking if the 'to_parse' and 'parsed_files'
#           folders exist
# shutil - used to create a copy of the template file to append.
# from datetime import datetime - used to get the month and year
# from openpyxl import load_workbook - used to read, modify and save the results in the .xlsx format

def start_parse():

    # Runs script through each step of the parsing process

    dir_exist = dir_check()

    if dir_exist:

        findings = find_csv_names()

        pulled_data = pull_data(findings)

        generated_report = generate_report(pulled_data)

        if generated_report:

            print('Vacation hours successfully updated')


        
def dir_check():

    # Will create necessary directories on first launch of script,
    # then will pass each other time. Gives a message to let the user know
    # what to do with the created directories

    if os.path.isdir('to_parse') and os.path.isdir('parsed_files'):

        return True

    else:

        create_these = ('to_parse', 'parsed_files')

        for i in create_these:

            if not os.path.isdir(i):

                os.makedirs(i)

        print('Folders for parsing created. Please place all files that need to be parsed',
              'into the "to_parse" folder, then run this script again.')

        return False
        

def find_csv_names():

    # Finds the names of the files to be parsed.
    
    return [file for file in os.listdir('to_parse') if file.endswith('.csv')]

def pull_data(names):

    # Takes the data from each CSV and loads it into memory to merged and written
    # to a single .xlsx file.

    # names - the file names of each file that needs to be parsed

    data = []

    for file in names:

        with open(f'to_parse/{file}', newline='') as csvfile:

            reader = csv.reader(csvfile, delimiter=',')

            next(reader, None)
            # Skips the header line

            for row in reader:

                first_name = row[0]
                last_name = row[1]
                hours_worked = row[4]

                vacation_earned = float(hours_worked) // 40
                # employees get 1 vacation hour for every 40 worked

                hours_remainder = round(float(hours_worked) % 40, 2)
                # banked hours that wi

                data.append((first_name, last_name, hours_remainder, vacation_earned))

    print(data)
                
    move_parsed_files(names)

    return data

def move_parsed_files(names):

    # Moves files that were parsed for the report to a new folder so the user knows they've been scanned

    # names - a list of file names that have been parsed

    for file in names:

        os.replace(f'to_parse/{file}',  f'parsed_files/{file}')

def generate_report(data):

    # Loads the master_vacation_hours.xlsx file, loads all the info from the csv, then combines this data
    # with what was already in the master_vacation_hours.xlsx file

    # data - table that contains tuples with all the information to go onto the report in the correct order
    #        that was created in the pull_data function.

    vacation_master = 'master_vacation_hours.xlsx'

    if os.path.exists(vacation_master):

        column_letters = ('A', 'B', 'C', 'D')
        # Letters of the columns used

        today = datetime.today()
        # used for building a file name

        year = str(today.year)[2:]
        # Only needs the last 2 digits of the year

        month = str(today.month - 1)

    else:

        print('master_vacation_hours.xlsx is missing. Please locate this file and place it in the same folder')
        print('as "vacation_hours_generator.py"!')

        return False
       
##        if int(month) == 0:
##
##            month = '12'
##            # Month 0 doesnt exist, it means december aka 12
##            
##            year = str(int(year) - 1)
##            # Changes the year too
##
##            # TO DO: make some corner case years work
##
##        elif len(month) == 1:
##
##            month = f'0{month}'
##            # wotc file formatting requires the month to have a 0 if necessary

    wb = load_workbook(filename=vacation_master)

    ws = wb.active

    last_row = int(ws.max_row)

    current_hours = {}

    if last_row > 1:

        for num in range(2, last_row + 1):

            current_row = []
            
            for column in column_letters:

                current_row.append(str((ws[f'{column}' + str(num)].value)))

            current_hours[current_row[0] + ' ' + current_row[1]] = [current_row[2], current_row[3]]

        for employee in data:

            name = employee[0] + ' ' + employee[1]
            remainder = employee[2]
            vaycay_earned = employee[3]

            if current_hours.get(name) != None:

                new_remainder_hours = float(current_hours[name][0]) + remainder

                if new_remainder_hours > 39:
                    
                    add_vaycay_from_remainder = new_remainder_hours // 40
                    # additonal vaycay earned from with addition of new remainder hours

                    new_remainder_hours = round(new_remainder_hours % 40, 2)
                    # New remainder after totaling old remainder and added remainder

                else:

                    add_vaycay_from_remainder = 0
                
                curr_vaycay_earned = int(current_hours[name][1]) + add_vaycay_from_remainder + vaycay_earned
                # New total vacation hours

                current_hours[name][0] = new_remainder_hours
                current_hours[name][1] = curr_vaycay_earned
                # Updates employee dict with new hours

            else:

                # If employee does not exist in system yet, create them
                current_hours[name] = [remainder, vaycay_earned]

                print(f'{name} has been added to the master vacation time spreadsheet')
            
    start_row = 2

    row_position = 0

    name_split = ''
        
    for row_num, row_data in enumerate(current_hours.items()):

        for column in column_letters:

            if row_position == 0:
                # first name

                name_split = row_data[0].split()

                ws[column + str(row_num + start_row)] = name_split[0]

            elif row_position == 1:
                # last name

                ws[column + str(row_num + start_row)] = name_split[1]

            elif row_position == 2:
                # remainder hours

                ws[column + str(row_num + start_row)] = row_data[1][0]

            else:
                # vacation hours

                ws[column + str(row_num + start_row)] = row_data[1][1]

            row_position += 1

        row_position = 0

    wb.save(vacation_master)

    return True
      

if __name__ == '__main__':

    start_parse()
